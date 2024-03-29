import openpyxl
import graphics

excel=openpyxl.load_workbook("pocha.xlsx")

lista_hojas=excel.sheetnames
data_puntos=[]
euros_totales=0
n_rondas_global=[]
lista_titles=[]
lista_ronda=[1,1,1,2,3,4,5,6,7,8,9,10,11,12,13,13,13,12,11,10,9,8,7,6,5,4,3,2,1,1,1]

class Jugador:
    
    def __init__(self,nombre):
        self.nombre=nombre
        self.n_vic=0 
        self.n_pos2=0
        self.n_pos3=0
        self.p_max=0
        self.p_min=0
        self.max_par=0
        self.min_par=0
        self.euros=0
        self.data_partida={}
        self.total_partida=[]
        self.acumulado=[]

    def carga_data_jug(self,partida,puntos):
        self.data_partida[partida]=puntos
    
    def set_n_vic(self,n):
        self.n_vic+=n
    def set_n_pos2(self,n):
        self.n_pos2+=n
    def set_n_pos3(self,n):
        self.n_pos3+=n
    def set_parcial_max(self,n):
        self.p_max=n
    def set_parcial_min(self,n):
        self.p_min=n
    def set_max_partida(self,n):
        self.max_par=n
    def set_min_partida(self,n):
        self.min_par=n
    def set_total_partida(self,lista):
        self.total_partida=lista
    def set_acumulado(self,lista):
        self.acumulado.append(lista)
    def set_euros(self):
        global euros_totales
        self.euros=self.n_pos2*3+self.n_pos3*5
        euros_totales+=self.euros
    def print_jugador(self):
        global euros_totales
        result=""
        result+="Jugador:"+str(self.nombre)+'\n'
        result+="Maximo partida:"+str(self.max_par)+" Minimo partida:"+str(self.min_par)+'\n'
        result+="Maximo parcial:"+str(self.p_max)+" Minimo parcial:"+str(self.p_min)+'\n'
        result+="Victorias:"+str(self.n_vic)+", 2ª Posicion:"+str(self.n_pos2)+", 3ª Posicion:"+str(self.n_pos3)+'\n'
        result+="Euros aportados:"+str(self.euros)+ "de un total de:"+str(euros_totales)+"->"+str(round((self.euros/euros_totales)*100,2))+"%"+'\n'
        return result

        
Teo=Jugador("T")
Visi=Jugador("V")
Alberto=Jugador("A")
lista_jugadores=[Teo,Visi,Alberto]

def carga_excel(jugador):
    global lista_hojas
    global data_puntos
    for partida in lista_hojas:
        data_partida=excel[partida]
        for columna in data_partida.iter_cols(values_only=True):
            for celda in columna:
                if celda == "T" or celda == "V" or celda == "A":
                    jugador_act=celda
                if not isinstance(celda,str):
                    data_puntos.append(celda)
                    if jugador_act==jugador.nombre:
                        jugador.carga_data_jug(partida,data_puntos.copy())
            data_puntos.clear()

def stats():
    global lista_jugadores
    for jugador in lista_jugadores:
        create_stats(jugador)
    victories_counter()
    for jugador in lista_jugadores:
        jugador.set_euros()
    
    
            

def create_stats(jugador):
    data=jugador.data_partida
    lista_maximos=[]
    lista_minimos=[]
    lista_suma=[]
    total=0
    for partida in data:
        max_=max(data[partida])
        min_=min(data[partida])
        for puntos in data[partida]:
            total+=puntos
        lista_suma.append(total)
        lista_maximos.append(max_)
        lista_minimos.append(min_)
        total=0
    jugador.set_parcial_max(max(lista_maximos))
    jugador.set_parcial_min(min(lista_minimos))
    jugador.set_max_partida(max(lista_suma))
    jugador.set_min_partida(min(lista_suma))
    jugador.set_total_partida(lista_suma)
    
def victories_counter():
    global lista_jugadores
    lista_todas_sumas=[]
    puntos_n_partida=[]
    for jugador in lista_jugadores:
        lista_todas_sumas.append(jugador.total_partida)
    for cont in range(len(lista_todas_sumas[0])):
        puntos=[sublista[cont] for sublista in lista_todas_sumas]
        puntos_n_partida.append(puntos.copy())
    for puntuaciones in puntos_n_partida:
        if max(puntuaciones) == puntuaciones[0]:
            lista_jugadores[0].set_n_vic(1)
            if puntuaciones[1]>puntuaciones[2]:
                lista_jugadores[1].set_n_pos2(1)
                lista_jugadores[2].set_n_pos3(1)
            else:
                lista_jugadores[2].set_n_pos2(1)
                lista_jugadores[1].set_n_pos3(1)
        if max(puntuaciones) == puntuaciones[1]:
            lista_jugadores[1].set_n_vic(1)
            if puntuaciones[0]>puntuaciones[2]:
                lista_jugadores[0].set_n_pos2(1)
                lista_jugadores[2].set_n_pos3(1)
            else:
                lista_jugadores[2].set_n_pos2(1)
                lista_jugadores[0].set_n_pos3(1)
        if max(puntuaciones) == puntuaciones[2]:
            lista_jugadores[2].set_n_vic(1)
            if puntuaciones[0]>puntuaciones[1]:
                lista_jugadores[0].set_n_pos2(1)
                lista_jugadores[1].set_n_pos3(1)
            else:
                lista_jugadores[1].set_n_pos2(1)
                lista_jugadores[0].set_n_pos3(1)

def pre_graphics_1p(nombre):
    global lista_jugadores
    global lista_titles
    n_rondas=[]
    lista_rondas=[]
    for jugador in lista_jugadores:
        load_acumulado(jugador)
    for cont in range(len(lista_jugadores[0].acumulado)):
        for i in range(len(lista_jugadores[0].acumulado[cont])):
            n_rondas.append(i)
        lista_rondas.append(n_rondas.copy())
        n_rondas.clear()
    return nombre,lista_rondas,lista_jugadores


def pre_graphics_multi():
    global lista_jugadores
    for jugador in lista_jugadores:
        load_acumulado(jugador)
    n_rondas=[]
    for cont in range(len(lista_jugadores[0].acumulado)):
        for i in range(len(lista_jugadores[0].acumulado[cont])):
            n_rondas.append(i)
        graphics.plot_multiplayer(n_rondas.copy(),lista_jugadores[0].acumulado[cont],lista_jugadores[1].acumulado[cont],lista_jugadores[2].acumulado[cont],str("Partida"+str(cont+1)))
        n_rondas.clear()

def load_acumulado(jugador):
    acumulado=[]
    n_rondas=[]
    global lista_titles
    for partida,puntos in jugador.data_partida.items():
        for cont,puntuacion in enumerate(puntos):
            if len(acumulado)==0:
                acumulado.append(puntuacion)
            else:
                acumulado.append(acumulado[cont-1]+puntuacion)
            n_rondas.append(cont)
        jugador.set_acumulado(acumulado.copy())
        acumulado.clear()
        n_rondas.clear()

def manage_punt():
    pass
        

def main():
    for jugador in lista_jugadores:
        carga_excel(jugador)
    stats()
    